import os
import re
from typing import List, Optional
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError, NoSuchTableError
# from openai import OpenAI
from pydantic import BaseModel
from config.config import Config
from config.logger import logger
from models.language_local_model import LanguageLocale
from models.language_strings_model import LanguageString
from datetime import datetime
import time
from openai import AsyncOpenAI
import asyncio
from openai import AsyncOpenAI
from typing import List, Tuple

# os.environ.pop("SSL_CERT_FILE", None) 

# client = OpenAI(api_key=Config.OPENAI_API_KEY)
client = AsyncOpenAI(api_key=Config.OPENAI_API_KEY)

class TranslationResult(BaseModel):
    translated_text: str


def get_zero_msgids(db: Session, lang_column: str) -> List[str]:
    try:
        logger.info(f"Getting zero msgids for: {lang_column}")
        if not hasattr(LanguageString, lang_column):
            error_msg = f"Invalid language column: {lang_column}"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        column_attr = getattr(LanguageString, lang_column)
        query = db.query(LanguageString.msgid).filter(column_attr == 0).order_by(LanguageString.id)
        return [row.msgid for row in query.all()]
    

    except NoSuchTableError:
        error_msg = f"Table '{LanguageString.__tablename__}' does not exist"
        logger.error(error_msg)
        raise RuntimeError(error_msg)
    except Exception as e:
        logger.error(f"Error getting zero msgids: {str(e)}")
        raise


async def translate_msgids(db: Session, msgids: List[str], lang_name: str, lang_code: str) -> List[str]:
    logger.info(f"Translating {len(msgids)} strings to {lang_name}")
    translations = []
    batch_size = 500  # Adjust based on your needs
    max_concurrency = 20  # Max parallel requests
    
    # Define the system prompt
    system_prompt = (
        "You are a professional POS translator. Translate text while EXACTLY preserving: "
        "• Punctuation, numbers, symbols, and formatting\n"
        "NEVER add/remove quotes or other characters.\n"
        "Return ONLY the translated text."
        "\nSTRICTLY DO NOT ADD ANY UNWANTED PUNCTUATION MARKS APPART FROM THE GIVEN"
    )

    async def translate_text(text: str, client: AsyncOpenAI, semaphore: asyncio.Semaphore) -> str:
        async with semaphore:
            response = await client.responses.parse(
                model="gpt-4o-2024-08-06",
                input=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"Translate the following to {lang_name} exactly as requested:\n\n{text}"}
                ],
                text_format=TranslationResult,
            )
            return response.output_parsed.translated_text

    async def process_batch(batch_msgids: List[str]) -> Tuple[List[str], List[str]]:
        """Translate a batch and return (translations, failed_msgids)"""
        semaphore = asyncio.Semaphore(max_concurrency)
        async with AsyncOpenAI() as client:
            tasks = [translate_text(msgid, client, semaphore) for msgid in batch_msgids]
            results = await asyncio.gather(*tasks, return_exceptions=True)
        
        batch_translations = []
        failed_msgids = []
        
        for i, result in enumerate(results):
            if isinstance(result, Exception):
                logger.error(f"Error translating '{batch_msgids[i]}': {str(result)}")
                batch_translations.append(batch_msgids[i])  # Use original on error
                failed_msgids.append(batch_msgids[i])
            else:
                batch_translations.append(result)
                print("###########@@@@@@@@@@@@@@@@@@@##############")
                print(batch_translations)
        
        # return batch_translations, failed_msgids
        return batch_translations
    # Process in batches
    for i in range(0, len(msgids), batch_size):
        batch = msgids[i:i + batch_size]
        logger.info(f"Processing batch {i//batch_size + 1}/{(len(msgids)-1)//batch_size + 1}")
        
        # batch_translations, failed_msgids = await process_batch(batch)
        batch_translations = await process_batch(batch)
        translations.extend(batch_translations)
        
        # Update database for successful translations
        success_msgids = set(batch)
        if success_msgids:
            # Fetch records in bulk
            records = db.query(LanguageString).filter(
                LanguageString.msgid.in_(list(success_msgids))
            ).all()
            
            # Create lookup dictionary
            record_dict = {r.msgid: r for r in records}
            not_found = []
            
            for msgid in success_msgids:
                if msgid in record_dict:
                    setattr(record_dict[msgid], lang_code, 1)
                else:
                    not_found.append(msgid)
            
            # Commit bulk updates
            try:
                db.commit()
                for msgid in not_found:
                    logger.warning(f"No record found for msgid: '{msgid}'")
            except Exception as e:
                db.rollback()
                logger.error(f"Database commit failed: {str(e)}")
    
    return translations

#######################################################################################
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from sqlalchemy.orm import Session
import csv
import os
from io import BytesIO, StringIO
import logging
from typing import List

def read_csv_file(content: bytes):
    """Read CSV file with encoding detection"""
    # Try common encodings in order
    encodings = ['utf-8', 'latin-1', 'cp1252', 'utf-16']
    decoded_content = None
    
    for encoding in encodings:
        try:
            decoded_content = content.decode(encoding)
            logger.info(f"Successfully decoded CSV using {encoding} encoding")
            break
        except UnicodeDecodeError:
            continue
    
    if decoded_content is None:
        raise HTTPException(
            status_code=400,
            detail="Unable to decode file. Supported encodings: UTF-8, Latin-1, CP1252, UTF-16"
        )
    
    return decoded_content

def read_excel_file(content: bytes, file_extension: str):
    """Read Excel file (XLSX or XLS)"""
    try:
        if file_extension == 'xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(filename=BytesIO(content))
            sheet = wb.active
            rows = sheet.iter_rows(values_only=True)
            return list(rows)
        
        elif file_extension == 'xls':
            import xlrd
            book = xlrd.open_workbook(file_contents=content)
            sheet = book.sheet_by_index(0)
            return [sheet.row_values(row) for row in range(sheet.nrows)]
        
    except ImportError:
        logger.error("Excel library not installed")
        raise HTTPException(
            status_code=500,
            detail="Excel processing requires openpyxl for XLSX or xlrd for XLS files"
        )
    except Exception as e:
        logger.error(f"Excel read error: {str(e)}")
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Excel file: {str(e)}"
        )

def extract_translations_from_rows(rows: list):
    """Extract translations from rows (CSV or Excel)"""
    if len(rows) < 1:
        raise HTTPException(
            status_code=400,
            detail="File is empty"
        )
    
    # Get header row
    headers = [str(cell).lower() if cell else "" for cell in rows[0]]
    
    # Find required columns
    try:
        msgid_index = headers.index("msgid")
        msgstr_index = headers.index("msgstr")
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="File must contain 'msgid' and 'msgstr' columns"
        )
    
    # Process data rows
    msgids = []
    translations = []
    line_number = 1  # Start after header
    
    for row in rows[1:]:
        line_number += 1
        try:
            # Handle different row formats (list vs tuple)
            row_data = list(row) if isinstance(row, tuple) else row
            
            # Ensure row has enough columns
            if len(row_data) <= max(msgid_index, msgstr_index):
                raise ValueError(f"Missing values in row {line_number}")
            
            msgid = str(row_data[msgid_index]) if row_data[msgid_index] is not None else ""
            msgstr = str(row_data[msgstr_index]) if row_data[msgstr_index] is not None else ""
            
            msgids.append(msgid)
            translations.append(msgstr)
            
        except Exception as e:
            logger.error(f"Error processing row {line_number}: {str(e)}")
            raise HTTPException(
                status_code=400,
                detail=f"Error in row {line_number}: {str(e)}"
            )
    
    return msgids, translations

def process_uploaded_translations(db: Session, lang_code: str, msgids: List[str], translations: List[str]):
    """
    Update database with uploaded translations
    """
    # Verify equal length
    if len(msgids) != len(translations):
        raise ValueError("msgids and translations lists must have the same length")
    
    records = db.query(LanguageString).filter(
        LanguageString.msgid.in_(msgids)
    ).all()
    
    record_dict = {r.msgid: r for r in records}
    not_found = []
    
    # Create new entries for missing msgids
    for i, msgid in enumerate(msgids):
        if msgid in record_dict:
            # Update existing translation
            setattr(record_dict[msgid], f"translation_{lang_code}", translations[i])
        else:
            # Create new record
            new_record = LanguageString(
                msgid=msgid,
                **{f"translation_{lang_code}": translations[i]}
            )
            db.add(new_record)
            not_found.append(msgid)
    
    try:
        db.commit()
        for msgid in not_found:
            logger.info(f"Created new record for msgid: '{msgid}'")
    except Exception as e:
        db.rollback()
        logger.error(f"Database commit failed: {str(e)}")
        raise



def get_language_code_by_name(db: Session, language: str) -> Optional[str]:
    try:
        logger.info(f"Getting language code for: {language}")
        record = db.query(LanguageLocale.language_code).filter(
            LanguageLocale.language == language
        ).first()
        return record[0] if record else None
    except SQLAlchemyError as e:
        logger.error(f"Database error: {str(e)}")
        raise

import zoneinfo
tz = zoneinfo.ZoneInfo("Asia/Colombo")
now = datetime.now(tz)
ts = now.strftime("%Y-%m-%d %H:%M%z")

def update_po_content(
    existing_content: str,
    lang_name: str,
    new_msgids: List[str],
    new_translations: List[str]
) -> str:
    # Get current date
    current_date = datetime.now().strftime("%Y-%m-%d")
    
    # Update the comment date
    existing_content = re.sub(
        r'(# date: )[\d-]+',
        f'\\g<1>{current_date}',
        existing_content
    )
    
    existing_content = re.sub(
        r'("PO-Revision-Date: )[\d:\s+-]+(?=\\n")',
        f'\\g<1>{ts}',
        existing_content
    )
    
    # Extract the body (after the header)
    header_end = existing_content.find('msgstr ""') + len('msgstr ""') + 1
    header = existing_content[:header_end]
    body = existing_content[header_end:]
    
    # Create lookup for existing translations
    existing_translations = {}
    for match in re.finditer(r'msgid "(.*?)"\nmsgstr "(.*?)"\n', body, re.DOTALL):
        msgid = match.group(1).replace('\\"', '"').replace('\\n', '\n')
        msgstr = match.group(2).replace('\\"', '"').replace('\\n', '\n')
        existing_translations[msgid] = msgstr
    
    # Ensure proper spacing at end of body
    if body and not body.endswith('\n\n'):
        body = body.rstrip() + '\n\n'
    
    # Add new translations (only for missing msgids)
    for msgid, translation in zip(new_msgids, new_translations):
        if msgid not in existing_translations:
            escaped_id = msgid.replace('"', '\\"').replace('\n', '\\n')
            escaped_tr = translation.replace('"', '\\"').replace('\n', '\\n')
            body += f'msgid "{escaped_id}"\n'
            body += f'msgstr "{escaped_tr}"\n\n'
    
    return header + body


# Original function remains for new file creation
def generate_po_content(lang_name: str,lang_code: str, msgids: List[str], translations: List[str]) -> str:
    logger.info("Generating PO content")
    content = [
        '# Autogenerated by SalesPlay Translate development',
        '#',
        f'# language: {lang_name}',
        f'# locale: {lang_code}',
        f'# date: {datetime.now().strftime("%Y-%m-%d")}',
        '#',
        'msgid ""',
        'msgstr ""',
        '"Project-Id-Version: SalesPlay POS Translation-0.000\\n"',
        f'"POT-Creation-Date: {ts}\\n"',
        f'"PO-Revision-Date: {ts}\\n"',
        '"Last-Translator: SalesPlay Team\\n"',
        '"Language-Team: SalesPlay (Pvt) Ltd <support@nvision.lk>\\n"',
        '"Language: en\\n"',
        '"MIME-Version: 1.0\\n"',
        '"Content-Type: text/plain; charset=UTF-8\\n"',
        '"Content-Transfer-Encoding: 8bit\\n"',
        '"Plural-Forms: nplurals=2; plural=n != 1;\\n"',
        '"X-Generator: Poedit 3.0.1\\n"',
        ''
    ]
    
    for msgid, msgstr in zip(msgids, translations):
        escaped_msgid = msgid.replace('"', '\\"').replace('\n', '\\n')
        escaped_msgstr = msgstr.replace('"', '\\"').replace('\n', '\\n')
        content.append(f'msgid "{escaped_msgid}"')
        content.append(f'msgstr "{escaped_msgstr}"\n')
    
    return "\n".join(content)